import openpyxl
import mysql.connector
import os
from os.path import dirname
#-----------------------------------------------------------------------------------------------
def open_Connection():
    return mysql.connector.connect(user='root', 
                              password='12345',
                              host='127.0.0.1',
                              port=3307,
                              database='Oraghe_Mosharekat',
                              auth_plugin='mysql_native_password')
#-----------------------------------------------------------------------------------------------
def add_Info(cnx, tableName, bran, amount):
    cursor = cnx.cursor()
    try:
        bran_info = """INSERT INTO %s (Bran, Amount) VALUES ('%s', '%s')""" %(tableName, str(bran), str(amount))
        cursor.execute(bran_info)
    except Exception as err:
        print(err)
#-----------------------------------------------------------------------------------------------
def loadDataFromFile():
    cnx = open_Connection()
    dir_path = os.path.dirname(os.path.realpath(__file__))
    dir_path = dirname(dir_path) + '\\files\input.xlsx'
    wookbook = openpyxl.load_workbook(dir_path)
    sheetnames = wookbook.sheetnames
    count = 0
    worksheet = wookbook[sheetnames[0]]
    for i in range(2, worksheet.max_row):
        add_Info(cnx, 'Report_202', worksheet.cell(row=i, column=1).value, worksheet.cell(row=i, column=2).value)
        count += 1
        if count % 5000 == 0:
            print('Report_202 - insert to DB' , count) 
            cnx.commit()
    print('Report_202 - insert to DB' , count) 
    cnx.commit()
    count = 0
    worksheet = wookbook[sheetnames[1]]
    for i in range(2, worksheet.max_row):
        add_Info(cnx, 'Total_GL', worksheet.cell(row=i, column=1).value, worksheet.cell(row=i, column=2).value)
        count += 1
        if count % 5000 == 0:
            print('Total_GL - insert to DB' , count) 
            cnx.commit()
    print('Total_GL - insert to DB' , count) 
    cnx.commit()
#-----------------------------------------------------------------------------------------------
def LoadData():
    loadDataFromFile()
#-----------------------------------------------------------------------------------------------
